"""Chuyển sheet danh sách đang làm việc của file HR sang JSON cho db-import-staff.ts.

Chạy: python3 scripts/xlsx-to-staff-json.py "<đường dẫn .xlsx>"

Đọc sheet thứ hai (283 người đang làm việc) làm nguồn chính. Sheet thứ nhất chỉ
dùng để tra số điện thoại — sheet thứ hai không có cột đó, mà cột `phone` của
bảng `users` không cho trống.
"""

import json
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

OUT = Path(__file__).parent / "data" / "staff-import.json"
HEADER_ROW = 4


def rows_of(ws):
    return [r for r in list(ws.iter_rows(values_only=True))[HEADER_ROW:] if r[0] not in (None, "")]


def no_accent(text: str) -> str:
    # 'đ' không tách được bằng NFD nên phải thay tay trước (AGENTS.md §9).
    text = text.replace("đ", "d").replace("Đ", "D")
    return "".join(c for c in unicodedata.normalize("NFD", text) if not unicodedata.combining(c))


def username_of(code: str) -> str:
    """Tên đăng nhập LÀ mã nhân viên viết thường — mã đã duy nhất nên không cần hậu tố.

    Schema `StaffForm` chỉ nhận `[a-z0-9._-]`, nên chữ hoa và dấu phải bỏ trước.
    """
    name = no_accent(code.strip()).lower()
    if not re.fullmatch(r"[a-z0-9._-]{3,}", name):
        raise SystemExit(f"Mã nhân viên không dùng làm tên đăng nhập được: {code}")
    return name


def phone_of(raw) -> str:
    if raw in (None, ""):
        return ""
    digits = re.sub(r"\D", "", str(raw))
    # Excel bỏ số 0 đầu khi cột lưu dạng số — 9 chữ số là số điện thoại đủ, thiếu 0.
    if len(digits) == 9:
        digits = "0" + digits
    return digits if re.fullmatch(r"0\d{9}", digits) else ""


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit("Thiếu đường dẫn file .xlsx")
    wb = load_workbook(sys.argv[1], data_only=True, read_only=True)
    sheets = wb.worksheets
    if len(sheets) < 2:
        raise SystemExit("File phải có ít nhất hai sheet")

    phone_by_code = {r[0]: phone_of(r[4]) for r in rows_of(sheets[0])}
    source = rows_of(sheets[1])

    # Mã trùng thì DỪNG HẲN — hai người cùng tên đăng nhập là một người mất hồ sơ.
    duplicated = [code for code, n in Counter(r[0].strip() for r in source).items() if n > 1]
    if duplicated:
        raise SystemExit(f"Mã nhân viên trùng trong file nguồn: {', '.join(duplicated)}")

    people = []
    for code, full_name, position, department, status in (r[:5] for r in source):
        people.append(
            {
                "staffCode": code.strip(),
                "username": username_of(code),
                "fullName": " ".join(str(full_name).split()),
                "position": str(position).strip(),
                "department": str(department).strip(),
                "status": str(status).strip(),
                "phone": phone_by_code.get(code, ""),
            }
        )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(people, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    no_phone = [p["staffCode"] for p in people if not p["phone"]]
    print(f"Ghi {len(people)} người vào {OUT}")
    print(f"Thiếu số điện thoại: {len(no_phone)} — {', '.join(no_phone) if no_phone else 'không có'}")


if __name__ == "__main__":
    main()
